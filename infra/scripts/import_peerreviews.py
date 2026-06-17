#!/usr/bin/env python3
"""
Import peer-review data from the Demokratiefabrik xlsx dumps into AT Protocol
as app.ch.poltr.peerreview.invitation and app.ch.poltr.peerreview.response records.

Reads two xlsx files:
  - content_peerreview.xlsx: 99 INSERT procedures (aggregated outcomes)
  - content_peerreview_progression.xlsx: 2,562 individual invitation/response rows

Records are written to a ballot-specific community account's PDS repo using
createRecord with composed rkeys ({content_id}-{mapped_did_suffix}), making
duplicates structurally impossible. Re-runs skip already-existing records.

Each ballot has its own community account (per-ballot community model).
Credentials are loaded from the community_accounts table in the DB.

Environment variables:
  PDS_HOST         PDS endpoint (default: http://localhost:2583)
  DB_URL           PostgreSQL connection URL (for community_accounts lookup)
  BALLOT_RKEY      Ballot rkey — credentials are loaded from community_accounts table
  MASTER_KEY_B64   APPVIEW_COMMUNITY_CREDS_MASTER_KEY_B64 (for decrypting community password)
  MAX_RESPONSES    Max number of responses to import (default: 0 = all)
  DRY_RUN          Set to "true" to inspect records without writing (default: false)
  PEERREVIEW_XLSX  Path to content_peerreview.xlsx (default: dump/content_peerreview.xlsx)
  PROGRESSION_XLSX Path to content_peerreview_progression.xlsx
                    (default: dump/content_peerreview_progression.xlsx)
"""

import base64
import os
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Optional

import openpyxl
import psycopg2
import requests
from nacl import secret as nacl_secret


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class PeerReview:
    id: int
    content_id: int
    operation: str
    date_created: Optional[str]


@dataclass
class Progression:
    id: int
    content_peerreview_id: int
    user_id: int
    response: Optional[int]  # 1=approve, 0=reject, None=no response
    date_created: Optional[str]
    date_responded: Optional[str]
    criteria_accept1: Optional[int]
    criteria_accept2: Optional[int]
    criteria_accept3: Optional[int]
    criteria_accept4: Optional[int]
    criteria_accept5: Optional[int]


@dataclass
class PdsUser:
    did: str
    handle: str


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

CRITERIA_MAP = [
    ("factual_accuracy", "Factual Accuracy"),
    ("relevance", "Relevance to Ballot"),
    ("clarity", "Clarity"),
    ("unity_of_thought", "Unity of Thought"),
    ("non_duplication", "Non-Duplication"),
]


def _compose_rkey(content_id: int, did: str) -> str:
    """Compose a deterministic rkey: {content_id}-{did_suffix}."""
    did_suffix = did.split(":")[-1]
    return f"{content_id}-{did_suffix}"


def _format_datetime(dt_str: Optional[str]) -> str:
    """Convert xlsx datetime string to ISO 8601."""
    if not dt_str:
        return datetime.now(timezone.utc).isoformat()
    try:
        if isinstance(dt_str, datetime):
            if dt_str.tzinfo is None:
                dt_str = dt_str.replace(tzinfo=timezone.utc)
            return dt_str.isoformat()
        dt = datetime.strptime(str(dt_str).strip(), "%Y-%m-%d %H:%M:%S")
        return dt.replace(tzinfo=timezone.utc).isoformat()
    except (ValueError, TypeError):
        return datetime.now(timezone.utc).isoformat()


def _criteria_to_rating(val: Optional[int]) -> int:
    """Map old binary criteria (0/1) to rating scale (1/5)."""
    if val is not None and int(val) == 1:
        return 5
    return 1


# ---------------------------------------------------------------------------
# DB credential loading
# ---------------------------------------------------------------------------


def load_community_creds(db_url: str, ballot_rkey: str, master_key_b64: str) -> tuple[str, str, str]:
    """Load community account DID, handle and decrypted password from DB.
    Returns (did, handle, password)."""
    conn = psycopg2.connect(db_url)
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT did, handle, pw_ciphertext, pw_nonce FROM auth.community_accounts WHERE ballot_rkey = %s",
                (ballot_rkey,),
            )
            row = cur.fetchone()
    finally:
        conn.close()

    if not row:
        print(f"ERROR: No community account found for ballot_rkey={ballot_rkey}")
        sys.exit(1)

    did, handle, pw_ct, pw_nonce = row
    pw_ct = bytes(pw_ct)
    pw_nonce = bytes(pw_nonce)

    key = base64.b64decode(master_key_b64)
    box = nacl_secret.SecretBox(key)
    password = box.decrypt(pw_ct, pw_nonce).decode("utf-8")

    return did, handle, password


# ---------------------------------------------------------------------------
# Importer
# ---------------------------------------------------------------------------

class PeerReviewImporter:
    def __init__(self, pds_host: str, community_handle: str, community_password: str,
                 ballot_uri: str, ballot_rkey: str,
                 dry_run: bool = False, max_responses: int = 0):
        self.pds_host = pds_host
        self.community_handle = community_handle
        self.community_password = community_password
        self.ballot_uri = ballot_uri
        self.ballot_rkey = ballot_rkey
        self.dry_run = dry_run
        self.max_responses = max_responses
        self.community_did: Optional[str] = None
        self.access_token: Optional[str] = None
        self.users: list[PdsUser] = []
        self.content_id_to_argument_uri: dict[int, str] = {}  # content_id -> AT URI

    def authenticate(self) -> bool:
        """Create a PDS session for the community account."""
        print(f"Authenticating as {self.community_handle}...")
        try:
            resp = requests.post(
                f"{self.pds_host}/xrpc/com.atproto.server.createSession",
                json={"identifier": self.community_handle, "password": self.community_password},
            )
            resp.raise_for_status()
            data = resp.json()
            self.access_token = data["accessJwt"]
            self.community_did = data["did"]
            print(f"  Authenticated as {self.community_did}")
            return True
        except Exception as e:
            print(f"ERROR: Authentication failed - {e}")
            return False

    def discover_users(self) -> bool:
        """List all repos on PDS to build the user pool for deterministic mapping."""
        print("Discovering PDS users...")
        repos_url = f"{self.pds_host}/xrpc/com.atproto.sync.listRepos?limit=1000"
        try:
            resp = requests.get(repos_url)
            resp.raise_for_status()
            repos = resp.json().get("repos", [])
        except Exception as e:
            print(f"ERROR: Failed to list repos - {e}")
            return False

        for repo in repos:
            did = repo["did"]
            try:
                desc = requests.get(
                    f"{self.pds_host}/xrpc/com.atproto.repo.describeRepo?repo={did}"
                )
                desc.raise_for_status()
                handle = desc.json().get("handle", "")
            except Exception:
                handle = did
            self.users.append(PdsUser(did=did, handle=handle))

        # Sort by DID for deterministic mapping
        self.users.sort(key=lambda u: u.did)
        print(f"  Found {len(self.users)} user(s) on PDS")
        for u in self.users:
            print(f"    {u.handle} ({u.did})")

        if not self.users:
            print("ERROR: No users found on PDS")
            return False
        return True

    def scan_existing_arguments(self):
        """Scan the community repo for argument records.
        Builds content_id → AT URI mapping using the rkey (which is the content_id).
        Only the community repo is scanned — arguments live exclusively there."""
        print(f"Scanning community repo ({self.community_did}) for argument records...")
        collection = "app.ch.poltr.ballot.argument"

        cursor = None
        while True:
            url = (
                f"{self.pds_host}/xrpc/com.atproto.repo.listRecords"
                f"?repo={self.community_did}&collection={collection}&limit=100"
            )
            if cursor:
                url += f"&cursor={cursor}"
            try:
                resp = requests.get(url)
                if resp.status_code != 200:
                    break
                data = resp.json()
                records = data.get("records", [])
                for rec in records:
                    uri = rec.get("uri", "")
                    rkey = uri.split("/")[-1]
                    ballot_ref = rec.get("value", {}).get("ballot", "")
                    # Accept both the new rkey-only format ("663.0") and the
                    # legacy AT-URI format ("at://…/app.ch.poltr.ballot.entry/663.0").
                    if rkey and (ballot_ref == self.ballot_rkey or ballot_ref == self.ballot_uri):
                        try:
                            content_id = int(rkey)
                            self.content_id_to_argument_uri[content_id] = uri
                        except ValueError:
                            pass
                cursor = data.get("cursor")
                if not cursor or not records:
                    break
            except Exception:
                break

        print(f"  Found {len(self.content_id_to_argument_uri)} argument(s) for this ballot")

    def _map_user_id_to_did(self, old_user_id: int) -> str:
        """Deterministic mapping: hash(old_user_id) % len(users)."""
        idx = hash(old_user_id) % len(self.users)
        return self.users[idx].did

    def _create_record(self, collection: str, rkey: str, record: dict) -> bool:
        """Write a record to the community PDS repo via createRecord.
        Fails if the rkey already exists (immutable — no updates)."""
        if self.dry_run:
            print(f"  [DRY RUN] createRecord {collection}/{rkey}")
            return True

        url = f"{self.pds_host}/xrpc/com.atproto.repo.createRecord"
        headers = {
            "Authorization": f"Bearer {self.access_token}",
            "Content-Type": "application/json",
        }
        payload = {
            "repo": self.community_did,
            "collection": collection,
            "rkey": rkey,
            "record": record,
        }

        try:
            time.sleep(0.05)
            resp = requests.post(url, headers=headers, json=payload)
            if resp.status_code in (200, 201):
                return True
            # PDS returns 500 InternalServerError when rkey already exists
            # with createRecord (no cleaner error code available)
            if resp.status_code == 500:
                print(f"  Skipped (already exists): {collection}/{rkey}")
                return True
            try:
                err = resp.json()
                msg = err.get("message") or err.get("error") or f"HTTP {resp.status_code}"
                print(f"  Failed ({resp.status_code}): {msg}")
            except Exception:
                print(f"  Failed ({resp.status_code}): {resp.text[:200]}")
            return False
        except requests.exceptions.RequestException as e:
            print(f"  Failed: {e}")
            return False

    def import_from_xlsx(self, peerreview_path: str, progression_path: str):
        """Read peer review data from xlsx files and import."""
        # --- Read peer reviews ---
        print(f"Reading peer reviews from: {peerreview_path}")
        wb = openpyxl.load_workbook(peerreview_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        peerreviews: dict[int, PeerReview] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = dict(zip(headers, row))
            if d.get("operation") != "INSERT":
                continue
            pr_id = int(d["id"])
            peerreviews[pr_id] = PeerReview(
                id=pr_id,
                content_id=int(d["content_id"]),
                operation=d["operation"],
                date_created=d.get("date_created"),
            )
        wb.close()
        print(f"  Found {len(peerreviews)} INSERT peer review procedure(s)")

        # --- Read progressions ---
        print(f"Reading progressions from: {progression_path}")
        wb = openpyxl.load_workbook(progression_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        progressions: list[Progression] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = dict(zip(headers, row))
            pr_id = int(d["content_peerreview_id"])
            if pr_id not in peerreviews:
                continue
            progressions.append(Progression(
                id=int(d["id"]),
                content_peerreview_id=pr_id,
                user_id=int(d["user_id"]),
                response=int(d["response"]) if d.get("response") is not None else None,
                date_created=d.get("date_created"),
                date_responded=d.get("date_responded"),
                criteria_accept1=int(d["criteria_accept1"]) if d.get("criteria_accept1") is not None else None,
                criteria_accept2=int(d["criteria_accept2"]) if d.get("criteria_accept2") is not None else None,
                criteria_accept3=int(d["criteria_accept3"]) if d.get("criteria_accept3") is not None else None,
                criteria_accept4=int(d["criteria_accept4"]) if d.get("criteria_accept4") is not None else None,
                criteria_accept5=int(d["criteria_accept5"]) if d.get("criteria_accept5") is not None else None,
            ))
        wb.close()
        print(f"  Found {len(progressions)} progression row(s) for INSERT procedures")

        # --- Group progressions by peer review ---
        pr_to_progressions: dict[int, list[Progression]] = {}
        for prog in progressions:
            pr_to_progressions.setdefault(prog.content_peerreview_id, []).append(prog)

        # --- Import ---
        invitations_created = 0
        responses_created = 0
        skipped_no_argument = 0
        failed = 0

        limit = self.max_responses if self.max_responses > 0 else None

        for pr_id, pr in sorted(peerreviews.items()):
            if limit and responses_created >= limit:
                print(f"\n  Reached MAX_RESPONSES={limit}, stopping")
                break

            argument_uri = self.content_id_to_argument_uri.get(pr.content_id)
            if not argument_uri:
                skipped_no_argument += 1
                continue

            progs = pr_to_progressions.get(pr_id, [])
            if not progs:
                continue

            arg_rkey = str(pr.content_id)
            print(f"\n  Peer review #{pr_id} for content_id={pr.content_id} ({len(progs)} progressions)")

            for prog in progs:
                if limit and responses_created >= limit:
                    break
                mapped_did = self._map_user_id_to_did(prog.user_id)
                did_suffix = mapped_did.split(":")[-1]
                rkey = f"{arg_rkey}-{did_suffix}"

                # Always create invitation
                invitation_record = {
                    "$type": "app.ch.poltr.peerreview.invitation",
                    "argument": argument_uri,
                    "invitee": mapped_did,
                    "createdAt": _format_datetime(prog.date_created),
                }

                if self._create_record("app.ch.poltr.peerreview.invitation", rkey, invitation_record):
                    invitations_created += 1
                else:
                    failed += 1
                    continue

                # Create response only if the reviewer actually responded
                if prog.response is not None:
                    vote = "APPROVE" if prog.response == 1 else "REJECT"
                    criteria = []
                    for i, (key, label) in enumerate(CRITERIA_MAP):
                        raw = getattr(prog, f"criteria_accept{i + 1}")
                        criteria.append({
                            "key": key,
                            "label": label,
                            "rating": _criteria_to_rating(raw),
                        })

                    response_record = {
                        "$type": "app.ch.poltr.peerreview.response",
                        "argument": argument_uri,
                        "reviewer": mapped_did,
                        "criteria": criteria,
                        "vote": vote,
                        "createdAt": _format_datetime(prog.date_responded),
                    }

                    if self._create_record("app.ch.poltr.peerreview.response", rkey, response_record):
                        responses_created += 1
                    else:
                        failed += 1

        print()
        print("=" * 50)
        print(f"Invitations created: {invitations_created}")
        print(f"Responses created:   {responses_created}")
        print(f"Skipped (no arg):    {skipped_no_argument}")
        print(f"Failed:              {failed}")
        if self.dry_run:
            print("(DRY RUN — no records were written)")
        print("=" * 50)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    pds_host = os.getenv("PDS_HOST", "http://localhost:2583")
    db_url = os.getenv("DB_URL", "")
    ballot_rkey = os.getenv("BALLOT_RKEY", "")
    master_key_b64 = os.getenv("MASTER_KEY_B64", "")
    max_responses = int(os.getenv("MAX_RESPONSES", "0"))
    dry_run = os.getenv("DRY_RUN", "false").lower() == "true"
    peerreview_path = os.getenv("PEERREVIEW_XLSX", "dump/content_peerreview.xlsx")
    progression_path = os.getenv("PROGRESSION_XLSX", "dump/content_peerreview_progression.xlsx")

    if not db_url or not ballot_rkey or not master_key_b64:
        print("ERROR: DB_URL, BALLOT_RKEY, and MASTER_KEY_B64 required")
        print("  DB_URL:          PostgreSQL connection URL")
        print("  BALLOT_RKEY:     Ballot rkey (e.g. '663')")
        print("  MASTER_KEY_B64:  APPVIEW_COMMUNITY_CREDS_MASTER_KEY_B64")
        sys.exit(1)

    # Load credentials from community_accounts table
    community_did, community_handle, community_password = load_community_creds(db_url, ballot_rkey, master_key_b64)
    ballot_uri = f"at://{community_did}/app.ch.poltr.ballot.entry/{ballot_rkey}"

    print("=== AT Protocol Peer Review Import ===")
    print(f"PDS Host:        {pds_host}")
    print(f"Community Account:     {community_handle} ({community_did})")
    print(f"Ballot URI:      {ballot_uri}")
    print(f"Max Responses:   {max_responses if max_responses > 0 else 'all'}")
    print(f"Dry Run:         {dry_run}")
    print(f"Peerreview XLSX: {peerreview_path}")
    print(f"Progression XLSX:{progression_path}")
    print()

    importer = PeerReviewImporter(pds_host, community_handle, community_password,
                                  ballot_uri, ballot_rkey, dry_run, max_responses)

    if not importer.authenticate():
        sys.exit(1)

    if not importer.discover_users():
        sys.exit(1)

    importer.scan_existing_arguments()

    print()
    importer.import_from_xlsx(peerreview_path, progression_path)


if __name__ == "__main__":
    main()
