#!/usr/bin/env python3
"""
Import PRO/CONTRA arguments from the Demokratiefabrik content.xlsx dump
into the community repo as app.ch.poltr.ballot.argument records.

All arguments are written to the community account's repo. The authorDid
field is set to a random non-admin user to simulate real authorship.

Environment variables:
  PDS_HOST                          PDS endpoint (default: http://localhost:2583)
  PDS_COMMUNITY_ACCOUNT_HANDLE     Community account handle
  PDS_COMMUNITY_ACCOUNT_PASSWORD   Community account password
  BALLOT_URI                        AT URI of the ballot to attach arguments to
  MAX_IMPORTS                       Number of arguments to import (default: 1, 0 = all)
  XLSX_PATH                         Path to content.xlsx (default: dump/content.xlsx)
  ADMIN_HANDLE                      Handle to exclude from author pool (default: admin.id.poltr.ch)
  INDEXER_POSTGRES_URL              Postgres connection string (reads auth.auth_creds for user DIDs)
"""

import os
import random
import sys
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Optional

import psycopg2
import openpyxl
import requests


@dataclass
class Argument:
    id: int
    title: str
    body: str
    type: str  # PRO or CONTRA
    date_created: Optional[str]


class ArgumentImporter:
    def __init__(self, pds_host: str, community_handle: str, community_password: str,
                 ballot_uri: str, admin_handle: str, db_url: str):
        self.pds_host = pds_host
        self.community_handle = community_handle
        self.community_password = community_password
        self.ballot_uri = ballot_uri
        self.admin_handle = admin_handle
        self.db_url = db_url
        self.community_did: Optional[str] = None
        self.community_token: Optional[str] = None
        self.author_dids: list[str] = []  # non-admin user DIDs for authorDid

    def authenticate_community(self) -> bool:
        """Create a PDS session for the community account."""
        print(f"Authenticating as community account: {self.community_handle}")
        try:
            resp = requests.post(
                f"{self.pds_host}/xrpc/com.atproto.server.createSession",
                json={"identifier": self.community_handle, "password": self.community_password},
            )
            resp.raise_for_status()
            data = resp.json()
            self.community_did = data["did"]
            self.community_token = data["accessJwt"]
            print(f"  Authenticated as {self.community_did}")
            return True
        except Exception as e:
            print(f"ERROR: Failed to authenticate community account - {e}")
            return False

    def load_author_dids(self):
        """Load non-admin user DIDs from auth.auth_creds for authorDid assignment."""
        print("Loading user DIDs from DB...")
        conn = psycopg2.connect(self.db_url)
        try:
            cur = conn.cursor()
            cur.execute("SELECT did, handle FROM auth.auth_creds")
            for did, handle in cur.fetchall():
                if handle == self.admin_handle:
                    print(f"  Skipping admin: {handle}")
                    continue
                if did == self.community_did:
                    print(f"  Skipping community: {handle}")
                    continue
                self.author_dids.append(did)
            print(f"  {len(self.author_dids)} user(s) available for author assignment")
        finally:
            conn.close()

    def create_argument(self, arg: Argument, author_did: str) -> bool:
        """Create an argument record in the community repo."""
        rkey = str(arg.id)

        record = {
            "$type": "app.ch.poltr.ballot.argument",
            "title": arg.title.strip(),
            "body": arg.body.strip(),
            "type": arg.type,
            "ballot": self.ballot_uri,
            # Origin language of title/body — keeps the record self-describing
            # (BCP-47). Imported content is Swiss German by default.
            "langs": [os.getenv("POLTR_DEFAULT_LANGUAGE", "de-CH")],
            "source": {
                "$type": "app.ch.poltr.ballot.argument#sourceUser",
                "authorDid": author_did,
            },
            "createdAt": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.%fZ")[:-4] + "Z",
        }

        payload = {
            "repo": self.community_did,
            "collection": "app.ch.poltr.ballot.argument",
            "rkey": rkey,
            "record": record,
        }

        try:
            time.sleep(0.05)
            response = requests.post(
                f"{self.pds_host}/xrpc/com.atproto.repo.putRecord",
                headers={
                    "Authorization": f"Bearer {self.community_token}",
                    "Content-Type": "application/json",
                },
                json=payload,
            )

            if response.status_code in (200, 201):
                data = response.json()
                if "uri" in data:
                    print(f"  [{arg.type:6s}] {arg.title[:50]} -> {data['uri']} (author: {author_did[-12:]})")
                    return True

            try:
                error_data = response.json()
                error_msg = error_data.get("message") or error_data.get("error") or f"HTTP {response.status_code}"
                print(f"  Failed ({response.status_code}): {error_msg}")
            except Exception:
                print(f"  Failed ({response.status_code}): {response.text[:200]}")

            return False

        except requests.exceptions.RequestException as e:
            print(f"  Failed: {e}")
            return False

    def import_from_xlsx(self, xlsx_path: str, max_imports: int = 1):
        """Read PRO/CONTRA arguments from xlsx and import to community repo."""
        print(f"Reading arguments from: {xlsx_path}")

        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        ws = wb[wb.sheetnames[0]]

        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        arguments = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            d = dict(zip(headers, row))
            if d.get("type") not in ("PRO", "CONTRA"):
                continue
            if d.get("deleted") == 1 or d.get("disabled") == 1:
                continue

            title = (d.get("title") or "").strip()
            body = (d.get("text") or "").strip()
            if not title or not body:
                continue

            arguments.append(Argument(
                id=d.get("id"),
                title=title,
                body=body,
                type=d.get("type"),
                date_created=str(d.get("date_created", "")),
            ))

        wb.close()
        print(f"Found {len(arguments)} valid arguments (PRO + CONTRA)")

        created = 0
        failed = 0
        limit = max_imports if max_imports > 0 else len(arguments)

        for arg in arguments[:limit]:
            author_did = random.choice(self.author_dids)
            if self.create_argument(arg, author_did):
                created += 1
            else:
                failed += 1

        print()
        print("=" * 42)
        print(f"Imported: {created}, Failed: {failed}")
        print("=" * 42)


def main():
    pds_host = os.getenv("PDS_HOST", "http://localhost:2583")
    community_handle = os.getenv("PDS_COMMUNITY_ACCOUNT_HANDLE", "")
    community_password = os.getenv("PDS_COMMUNITY_ACCOUNT_PASSWORD", "")
    ballot_uri = os.getenv("BALLOT_URI", "")
    max_imports = int(os.getenv("MAX_IMPORTS", "1"))
    xlsx_path = os.getenv("XLSX_PATH", "dump/content.xlsx")
    admin_handle = os.getenv("ADMIN_HANDLE", "admin.id.poltr.ch")
    db_url = os.getenv("INDEXER_POSTGRES_URL", "")

    if not community_handle or not community_password:
        print("ERROR: PDS_COMMUNITY_ACCOUNT_HANDLE and PDS_COMMUNITY_ACCOUNT_PASSWORD required")
        sys.exit(1)

    if not ballot_uri:
        print("ERROR: BALLOT_URI required (AT URI of the ballot to attach arguments to)")
        print("  e.g. BALLOT_URI='at://did:plc:.../app.ch.poltr.ballot.entry/663'")
        sys.exit(1)

    if not db_url:
        print("ERROR: INDEXER_POSTGRES_URL required (for loading user DIDs)")
        sys.exit(1)

    print("=== AT Protocol Argument Import (Community Repo) ===")
    print(f"PDS Host:    {pds_host}")
    print(f"Community:  {community_handle}")
    print(f"Ballot URI:  {ballot_uri}")
    print(f"Max imports: {max_imports if max_imports > 0 else 'all'}")
    print(f"XLSX path:   {xlsx_path}")
    print()

    importer = ArgumentImporter(pds_host, community_handle, community_password,
                                ballot_uri, admin_handle, db_url)

    if not importer.authenticate_community():
        sys.exit(1)

    importer.load_author_dids()

    if not importer.author_dids:
        print("ERROR: No non-admin users found for author assignment")
        sys.exit(1)

    print()
    importer.import_from_xlsx(xlsx_path, max_imports=max_imports)


if __name__ == "__main__":
    main()
