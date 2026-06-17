#!/usr/bin/env python3
"""
Import COMMENT entries from the Demokratiefabrik content.xlsx dump
into AT Protocol as app.ch.poltr.comment records.

Supports both root comments (parent_id references an argument) and nested
replies (parent_id references another comment).  Nested replies walk up the
parent chain to resolve the root argument URI, and set the `parent` field to
the direct parent comment's AT-URI.  Comments are topologically sorted so
parents are always created before their children.

Comments are assigned to random non-admin PDS users (excluding the parent
comment's author for nested replies).  Authenticates using stored app
passwords from the auth.auth_creds table (same as the indexer) so
credentials stay intact.

Environment variables:
  PDS_HOST                          PDS endpoint (default: http://localhost:2583)
  PDS_ADMIN_PASSWORD                PDS admin password (Basic auth, for user discovery)
  BALLOT_URI                        AT URI of the ballot to scope arguments to
  MAX_IMPORTS                       Number of comments to import (default: 1, 0 = all)
  XLSX_PATH                         Path to content.xlsx (default: dump/content.xlsx)
  ADMIN_HANDLE                      Handle to exclude (default: admin.id.poltr.ch)
  INDEXER_POSTGRES_URL              Postgres connection string (reads auth.auth_creds)
  APPVIEW_USER_CREDS_MASTER_KEY_B64  Base64-encoded NaCl secret key for app password decryption
"""

import base64
import os
import random
import sys
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Optional

import nacl.secret
import nacl.utils
import openpyxl
import psycopg2
import requests


COMMENT_BODY_MAX_LEN = 5000  # must match app.ch.poltr.comment lexicon body.maxLength


@dataclass
class Comment:
    id: int
    parent_id: int
    title: str
    body: str
    date_created: Optional[str]


@dataclass
class PdsUser:
    did: str
    handle: str
    access_token: Optional[str] = None


class CommentImporter:
    def __init__(self, pds_host: str, pds_admin_password: str, ballot_uri: str,
                 admin_handle: str, db_url: str, master_key_b64: str):
        self.pds_host = pds_host
        self.pds_admin_password = pds_admin_password
        self.ballot_uri = ballot_uri
        self.admin_handle = admin_handle
        self.db_url = db_url
        self.master_key = base64.b64decode(master_key_b64)
        self.users: list[PdsUser] = []
        self.all_users: list[PdsUser] = []
        self.argument_rkey_to_uri: dict[str, str] = {}  # argument rkey -> full AT URI
        self.existing_comment_rkey_to_did: dict[str, str] = {}  # comment rkey -> did
        self._app_passwords: dict[str, str] = {}  # did -> decrypted app password

    def _admin_auth_header(self) -> str:
        token = base64.b64encode(f"admin:{self.pds_admin_password}".encode()).decode()
        return f"Basic {token}"

    def _load_app_passwords(self):
        """Load and decrypt app passwords from auth.auth_creds."""
        print("Loading stored app passwords from DB...")
        conn = psycopg2.connect(self.db_url)
        try:
            cur = conn.cursor()
            cur.execute("SELECT did, app_pw_ciphertext, app_pw_nonce FROM auth.auth_creds")
            box = nacl.secret.SecretBox(self.master_key)
            for did, ciphertext, nonce in cur.fetchall():
                try:
                    plaintext = box.decrypt(bytes(ciphertext), bytes(nonce))
                    self._app_passwords[did] = plaintext.decode()
                except Exception as e:
                    print(f"  WARNING: Failed to decrypt app password for {did}: {e}")
            print(f"  Loaded {len(self._app_passwords)} app password(s)")
        finally:
            conn.close()

    def discover_users(self) -> bool:
        """List all repos on PDS, resolve handles, exclude admin."""
        print("Discovering PDS users...")

        repos_url = f"{self.pds_host}/xrpc/com.atproto.sync.listRepos?limit=1000"
        try:
            resp = requests.get(repos_url)
            resp.raise_for_status()
            repos = resp.json().get("repos", [])
        except Exception as e:
            print(f"ERROR: Failed to list repos - {e}")
            return False

        for repo in repos:
            did = repo["did"]
            try:
                desc = requests.get(f"{self.pds_host}/xrpc/com.atproto.repo.describeRepo?repo={did}")
                desc.raise_for_status()
                handle = desc.json().get("handle", "")
            except Exception:
                handle = did

            user = PdsUser(did=did, handle=handle)
            self.all_users.append(user)

            if handle == self.admin_handle:
                print(f"  Skipping admin: {handle}")
                continue

            self.users.append(user)
            print(f"  Found user: {handle} ({did})")

        if not self.users:
            print("ERROR: No non-admin users found on PDS")
            return False

        print(f"  {len(self.users)} users available for comment assignment")
        return True

    def scan_existing_arguments(self):
        """Scan all user repos for existing argument records.
        Builds rkey->URI mapping so comments can reference them."""
        print("Scanning PDS for existing argument records...")
        collection = "app.ch.poltr.ballot.argument"

        for user in self.all_users:
            cursor = None
            while True:
                url = (
                    f"{self.pds_host}/xrpc/com.atproto.repo.listRecords"
                    f"?repo={user.did}&collection={collection}&limit=100"
                )
                if cursor:
                    url += f"&cursor={cursor}"
                try:
                    resp = requests.get(url)
                    if resp.status_code != 200:
                        break
                    data = resp.json()
                    records = data.get("records", [])
                    for rec in records:
                        uri = rec.get("uri", "")
                        rkey = uri.split("/")[-1]
                        # Only include arguments belonging to this ballot
                        ballot_ref = rec.get("value", {}).get("ballot", "")
                        if rkey and ballot_ref == self.ballot_uri:
                            self.argument_rkey_to_uri[rkey] = uri
                    cursor = data.get("cursor")
                    if not cursor or not records:
                        break
                except Exception:
                    break

        print(f"  Found {len(self.argument_rkey_to_uri)} argument(s) for this ballot")

    def scan_existing_comments(self):
        """Scan all user repos for existing comment records.
        Builds rkey->did mapping so re-imports reuse the same account."""
        print("Scanning PDS for existing comment records...")
        collection = "app.ch.poltr.comment"

        for user in self.all_users:
            cursor = None
            while True:
                url = (
                    f"{self.pds_host}/xrpc/com.atproto.repo.listRecords"
                    f"?repo={user.did}&collection={collection}&limit=100"
                )
                if cursor:
                    url += f"&cursor={cursor}"
                try:
                    resp = requests.get(url)
                    if resp.status_code != 200:
                        break
                    data = resp.json()
                    records = data.get("records", [])
                    for rec in records:
                        rkey = rec.get("uri", "").split("/")[-1]
                        if rkey:
                            self.existing_comment_rkey_to_did[rkey] = user.did
                    cursor = data.get("cursor")
                    if not cursor or not records:
                        break
                except Exception:
                    break

        if self.existing_comment_rkey_to_did:
            print(f"  Found {len(self.existing_comment_rkey_to_did)} existing comment(s) on PDS")
        else:
            print("  No existing comments found on PDS")

    def _get_user_by_did(self, did: str) -> Optional[PdsUser]:
        """Find a user by DID from all_users."""
        return next((u for u in self.all_users if u.did == did), None)

    def authenticate_user(self, user: PdsUser) -> bool:
        """Create session using stored app password from auth.auth_creds."""
        if user.access_token:
            return True

        password = self._app_passwords.get(user.did)
        if not password:
            print(f"  ERROR: No stored app password for {user.handle}")
            return False

        try:
            resp = requests.post(
                f"{self.pds_host}/xrpc/com.atproto.server.createSession",
                json={"identifier": user.did, "password": password},
            )
            resp.raise_for_status()
            data = resp.json()
            user.access_token = data.get("accessJwt")
            return bool(user.access_token)
        except Exception as e:
            print(f"  ERROR: Failed to create session for {user.handle} - {e}")
            return False

    def create_comment(self, comment: Comment, user: PdsUser, argument_uri: str,
                       parent_uri: Optional[str] = None) -> Optional[str]:
        """Create or update a comment record on the PDS as the given user.
        Returns the AT-URI of the created record, or None on failure."""
        rkey = str(comment.id)

        body_text = comment.body.strip()
        if len(body_text) > COMMENT_BODY_MAX_LEN:
            print(f"  Trimmed body from {len(body_text)} to {COMMENT_BODY_MAX_LEN} chars (id={comment.id})")
            body_text = body_text[:COMMENT_BODY_MAX_LEN]

        record = {
            "$type": "app.ch.poltr.comment",
            "title": comment.title.strip(),
            "body": body_text,
            "argument": argument_uri,
            # Origin language of the body (BCP-47) — imported comments are Swiss
            # German by default. Keeps the record self-describing.
            "langs": [os.getenv("POLTR_DEFAULT_LANGUAGE", "de-CH")],
            "createdAt": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.%fZ")[:-4] + "Z",
        }
        if parent_uri:
            record["parent"] = parent_uri

        url = f"{self.pds_host}/xrpc/com.atproto.repo.putRecord"
        headers = {
            "Authorization": f"Bearer {user.access_token}",
            "Content-Type": "application/json",
        }
        payload = {
            "repo": user.did,
            "collection": "app.ch.poltr.comment",
            "rkey": rkey,
            "record": record,
        }

        try:
            time.sleep(0.05)
            response = requests.post(url, headers=headers, json=payload)

            if response.status_code in (200, 201):
                data = response.json()
                if "uri" in data:
                    arg_rkey = argument_uri.split("/")[-1]
                    parent_info = f" (parent:{parent_uri.split('/')[-1]})" if parent_uri else ""
                    print(f"  Synced by {user.handle}: {comment.title[:40]} -> {data['uri']} (arg:{arg_rkey}){parent_info}")
                    return data["uri"]

            try:
                error_data = response.json()
                error_msg = error_data.get("message") or error_data.get("error") or f"HTTP {response.status_code}"
                print(f"  Failed ({response.status_code}): {error_msg}")
            except Exception:
                print(f"  Failed ({response.status_code}): {response.text[:200]}")

            return None

        except requests.exceptions.RequestException as e:
            print(f"  Failed: {e}")
            return None

    def _resolve_argument_uri(self, comment_id: int, all_comments: dict[int, Comment]) -> Optional[str]:
        """Walk up the parent_id chain from a comment until an argument is found.
        Returns the argument AT-URI, or None if the chain is broken."""
        visited: set[int] = set()
        current = comment_id
        while current in all_comments:
            if current in visited:
                return None  # cycle detected
            visited.add(current)
            parent_id = all_comments[current].parent_id
            parent_rkey = str(parent_id)
            if parent_rkey in self.argument_rkey_to_uri:
                return self.argument_rkey_to_uri[parent_rkey]
            current = parent_id
        return None

    def _topological_sort(self, comment_ids: list[int], all_comments: dict[int, Comment]) -> list[int]:
        """Sort comments so parents come before children."""
        comment_id_set = set(comment_ids)
        # Build adjacency: parent -> children
        children: dict[int, list[int]] = {cid: [] for cid in comment_ids}
        in_degree: dict[int, int] = {cid: 0 for cid in comment_ids}
        for cid in comment_ids:
            pid = all_comments[cid].parent_id
            if pid in comment_id_set:
                children[pid].append(cid)
                in_degree[cid] += 1
        # Kahn's algorithm
        queue = [cid for cid in comment_ids if in_degree[cid] == 0]
        result = []
        while queue:
            node = queue.pop(0)
            result.append(node)
            for child in children[node]:
                in_degree[child] -= 1
                if in_degree[child] == 0:
                    queue.append(child)
        return result

    def import_from_xlsx(self, xlsx_path: str, max_imports: int = 1):
        """Read COMMENT entries from xlsx and import them as random users."""
        print(f"Reading comments from: {xlsx_path}")

        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        ws = wb[wb.sheetnames[0]]

        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        # Pass 1: Read ALL comment rows into a dict keyed by id
        all_comments: dict[int, Comment] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = dict(zip(headers, row))
            if d.get("type") != "COMMENT":
                continue
            if d.get("contenttree_id") == 5:
                continue  # skip INVITED_COMMENT tree (ballot-level opinions, not argument replies)
            if d.get("deleted") == 1 or d.get("disabled") == 1:
                continue

            parent_id = d.get("parent_id")
            if parent_id is None:
                continue

            title = (d.get("title") or "").strip()
            body = (d.get("text") or "").strip()
            if not body:
                continue

            cid = int(d.get("id"))
            all_comments[cid] = Comment(
                id=cid,
                parent_id=int(parent_id),
                title=title,
                body=body,
                date_created=str(d.get("date_created", "")),
            )

        wb.close()

        # Pass 2: Classify each comment
        root_comments: list[int] = []      # parent_id -> argument
        nested_comments: list[int] = []    # parent_id -> another comment
        skipped = 0

        for cid, comment in all_comments.items():
            parent_rkey = str(comment.parent_id)
            if parent_rkey in self.argument_rkey_to_uri:
                root_comments.append(cid)
            elif comment.parent_id in all_comments:
                nested_comments.append(cid)
            else:
                skipped += 1

        # For nested comments, resolve the argument URI by walking up the chain
        # Filter out any whose chain doesn't reach an argument
        valid_nested: list[int] = []
        for cid in nested_comments:
            arg_uri = self._resolve_argument_uri(cid, all_comments)
            if arg_uri:
                valid_nested.append(cid)
            else:
                skipped += 1

        all_importable = root_comments + valid_nested
        print(f"Found {len(root_comments)} root comment(s), {len(valid_nested)} nested reply/replies, {skipped} orphan(s) skipped")

        # Topological sort so parents are created before children
        sorted_ids = self._topological_sort(all_importable, all_comments)

        created = 0
        failed = 0
        limit = max_imports if max_imports > 0 else len(sorted_ids)

        # Track created comment URIs and assigned authors
        comment_id_to_uri: dict[int, str] = {}
        comment_id_to_did: dict[int, str] = {}

        for cid in sorted_ids[:limit]:
            comment = all_comments[cid]
            rkey = str(comment.id)
            parent_rkey = str(comment.parent_id)

            # Determine argument URI and parent URI
            if parent_rkey in self.argument_rkey_to_uri:
                argument_uri = self.argument_rkey_to_uri[parent_rkey]
                parent_uri = None
            else:
                argument_uri = self._resolve_argument_uri(cid, all_comments)
                parent_uri = comment_id_to_uri.get(comment.parent_id)
                if not parent_uri:
                    print(f"  WARNING: Parent comment {comment.parent_id} not yet created for comment {cid}, skipping")
                    failed += 1
                    continue

            # Pick user: reuse existing, or random (excluding parent's author for nested)
            existing_did = self.existing_comment_rkey_to_did.get(rkey)
            if existing_did:
                user = self._get_user_by_did(existing_did)
                if not user:
                    print(f"  WARNING: existing DID {existing_did} for rkey={rkey} not found, picking random")
                    user = random.choice(self.users)
                else:
                    print(f"  Reusing existing account {user.handle} for rkey={rkey}")
            else:
                parent_author_did = comment_id_to_did.get(comment.parent_id)
                candidates = [u for u in self.users if u.did != parent_author_did] if parent_author_did else self.users
                if not candidates:
                    candidates = self.users  # fallback if only one user
                user = random.choice(candidates)

            if not self.authenticate_user(user):
                failed += 1
                continue

            uri = self.create_comment(comment, user, argument_uri, parent_uri=parent_uri)
            if uri:
                comment_id_to_uri[cid] = uri
                comment_id_to_did[cid] = user.did
                created += 1
            else:
                failed += 1

        print()
        print("=" * 42)
        print(f"Imported: {created} ({len(root_comments)} root + {created - len([c for c in root_comments if c in comment_id_to_uri]) if created else 0} nested), Failed: {failed}")
        print("=" * 42)


def main():
    pds_host = os.getenv("PDS_HOST", "http://localhost:2583")
    pds_admin_password = os.getenv("PDS_ADMIN_PASSWORD", "")
    ballot_uri = os.getenv("BALLOT_URI", "")
    max_imports = int(os.getenv("MAX_IMPORTS", "1"))
    xlsx_path = os.getenv("XLSX_PATH", "dump/content.xlsx")
    admin_handle = os.getenv("ADMIN_HANDLE", "admin.id.poltr.ch")
    db_url = os.getenv("INDEXER_POSTGRES_URL", "")
    master_key_b64 = os.getenv("APPVIEW_USER_CREDS_MASTER_KEY_B64", "")

    if not pds_admin_password:
        print("ERROR: PDS_ADMIN_PASSWORD required")
        sys.exit(1)

    if not ballot_uri:
        print("ERROR: BALLOT_URI required (AT URI of the ballot)")
        print("  e.g. BALLOT_URI='at://did:plc:.../app.ch.poltr.ballot.entry/663'")
        sys.exit(1)

    if not db_url or not master_key_b64:
        print("ERROR: INDEXER_POSTGRES_URL and APPVIEW_USER_CREDS_MASTER_KEY_B64 required")
        print("  (used to read stored app passwords from auth.auth_creds)")
        sys.exit(1)

    print("=== AT Protocol Comment Import ===")
    print(f"PDS Host:    {pds_host}")
    print(f"Ballot URI:  {ballot_uri}")
    print(f"Max imports: {max_imports if max_imports > 0 else 'all'}")
    print(f"XLSX path:   {xlsx_path}")
    print()

    importer = CommentImporter(pds_host, pds_admin_password, ballot_uri,
                               admin_handle, db_url, master_key_b64)

    if not importer.discover_users():
        sys.exit(1)

    importer._load_app_passwords()
    importer.scan_existing_arguments()
    importer.scan_existing_comments()

    print()
    importer.import_from_xlsx(xlsx_path, max_imports=max_imports)


if __name__ == "__main__":
    main()
